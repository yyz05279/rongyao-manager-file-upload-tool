#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
项目日报Excel解析工具
功能：解析项目日报Excel文件，生成可保存到数据库的JSON数据
"""

import openpyxl
import json
import sys
from datetime import datetime
from typing import Dict, List, Any

class DailyReportExcelParser:
    """日报Excel解析器"""
    
    def __init__(self, excel_path: str):
        """
        初始化解析器
        :param excel_path: Excel文件路径
        """
        self.excel_path = excel_path
        self.workbook = openpyxl.load_workbook(excel_path)
        
    def parse_sheet(self, sheet_name: str = None) -> Dict[str, Any]:
        """
        解析指定工作表
        :param sheet_name: 工作表名称，不指定则使用活动工作表
        :return: 解析后的数据字典
        """
        if sheet_name:
            ws = self.workbook[sheet_name]
        else:
            ws = self.workbook.active
            
        report_data = {
            "reportDate": sheet_name if sheet_name else ws.title,
            "reporterName": self._get_cell_value(ws, 1, 1).replace("项目工作日报", "").strip(),  # ✅ 改为 reporterName
            "overallProgress": None,
            "progressDescription": None,
            "taskProgressList": [],
            "tomorrowPlans": [],
            "workerReports": [],
            "machineryRentals": [],
            "problemFeedbacks": [],
            "requirements": [],
            "weather": None,
            "temperature": None,
            "onSitePersonnelCount": 0,
            "remarks": None
        }
        
        # 解析项目整体进度 (第3行)
        progress_desc = self._get_cell_value(ws, 3, 5)
        report_data["progressDescription"] = progress_desc
        # 根据描述内容判断进度状态
        if "正常" in progress_desc:
            report_data["overallProgress"] = "normal"
        elif "滞后" in progress_desc:
            report_data["overallProgress"] = "delayed"
        elif "超前" in progress_desc:
            report_data["overallProgress"] = "ahead"
        else:
            report_data["overallProgress"] = "normal"
        
        # 解析逐项进度汇报 (第6行开始，只保留序号2.x)
        report_data["taskProgressList"] = self._parse_task_progress(ws, start_row=6, end_row=20)
        
        # 解析明天工作计划 (第6行开始，只保留序号3.x)
        report_data["tomorrowPlans"] = self._parse_tomorrow_plans(ws, start_row=6, end_row=20)
        
        # 解析各工种工作汇报 (从第20行开始扫描，自动检测"二"区域)
        report_data["workerReports"] = self._parse_worker_reports(ws, start_row=20, end_row=70)
        
        # 统计现场总人数
        report_data["onSitePersonnelCount"] = len([w for w in report_data["workerReports"] if w.get("name")])
        
        # 解析机械租赁情况 (从第20行开始扫描，自动检测"三"区域)
        report_data["machineryRentals"] = self._parse_machinery_rentals(ws, start_row=20, end_row=70)
        
        # 解析问题反馈 (从第20行开始扫描，自动检测"四"区域，只保留序号1.x)
        report_data["problemFeedbacks"] = self._parse_problem_feedbacks(ws, start_row=20, end_row=80)
        
        # 解析需求描述 (从第20行开始扫描，自动检测"四"区域，只保留序号2.x)
        report_data["requirements"] = self._parse_requirements(ws, start_row=20, end_row=80)
        
        return report_data
    
    def _get_cell_value(self, ws, row: int, col: int) -> str:
        """获取单元格值，返回字符串"""
        value = ws.cell(row=row, column=col).value
        return str(value).strip() if value is not None else ""
    
    def _parse_task_progress(self, ws, start_row: int, end_row: int) -> List[Dict]:
        """解析逐项进度汇报（序号2.x）"""
        tasks = []
        for i in range(start_row, end_row + 1):
            task_no = self._get_cell_value(ws, i, 1)
            task_name = self._get_cell_value(ws, i, 2)
            planned_progress = self._get_cell_value(ws, i, 3)
            actual_progress = self._get_cell_value(ws, i, 5)
            deviation_reason = self._get_cell_value(ws, i, 6)
            impact_measures = self._get_cell_value(ws, i, 7)
            
            # 只保存有内容且序号以"2."开头的任务
            if task_name and task_no.startswith("2."):
                tasks.append({
                    "taskNo": task_no,
                    "taskName": task_name,
                    "plannedProgress": planned_progress,
                    "actualProgress": actual_progress,
                    "deviationReason": deviation_reason,
                    "impactMeasures": impact_measures
                })
        return tasks
    
    def _parse_tomorrow_plans(self, ws, start_row: int, end_row: int) -> List[Dict]:
        """解析明天工作计划（序号3.x）"""
        plans = []
        for i in range(start_row, end_row + 1):
            plan_no = self._get_cell_value(ws, i, 1)
            task_name = self._get_cell_value(ws, i, 2)
            goal = self._get_cell_value(ws, i, 3)
            responsible_person = self._get_cell_value(ws, i, 5)
            required_resources = self._get_cell_value(ws, i, 6)
            remarks = self._get_cell_value(ws, i, 7)
            
            # 只保存有内容且序号以"3."开头的计划
            if task_name and plan_no.startswith("3."):
                plans.append({
                    "planNo": plan_no,
                    "taskName": task_name,
                    "goal": goal,
                    "responsiblePerson": responsible_person,
                    "requiredResources": required_resources,
                    "remarks": remarks
                })
        return plans
    
    def _parse_worker_reports(self, ws, start_row: int, end_row: int) -> List[Dict]:
        """解析各工种工作汇报（区域二）"""
        workers = []
        in_target_area = False  # 是否进入目标区域（二）
        
        for i in range(start_row, end_row + 1):
            seq_no = self._get_cell_value(ws, i, 1)
            name = self._get_cell_value(ws, i, 2)
            
            # 检测区域标题
            if seq_no == '二':
                in_target_area = True
                continue  # 跳过标题行
            
            # 遇到下一个区域标题，停止解析
            if seq_no in ['三', '四', '五'] and in_target_area:
                break
            
            # 只在目标区域内解析数据
            if not in_target_area:
                continue
            
            # 跳过表头行
            if name in ['姓名', '序号']:
                continue
            
            job_type = self._get_cell_value(ws, i, 3)
            worker_type = self._get_cell_value(ws, i, 4)
            work_content = self._get_cell_value(ws, i, 5)
            work_hours = self._get_cell_value(ws, i, 7)
            
            # 只保存有姓名的记录
            if name:
                workers.append({
                    "seqNo": seq_no,
                    "name": name,
                    "jobType": job_type,
                    "workerType": worker_type,
                    "workContent": work_content,
                    "workHours": work_hours
                })
        return workers
    
    def _parse_machinery_rentals(self, ws, start_row: int, end_row: int) -> List[Dict]:
        """解析机械租赁情况（区域三）"""
        machinery = []
        in_target_area = False  # 是否进入目标区域（三）
        
        for i in range(start_row, end_row + 1):
            seq_no = self._get_cell_value(ws, i, 1)
            machine_name = self._get_cell_value(ws, i, 2)
            
            # 检测区域标题
            if seq_no == '三':
                in_target_area = True
                continue  # 跳过标题行
            
            # 遇到下一个区域标题，停止解析
            if seq_no in ['四', '五', '六'] and in_target_area:
                break
            
            # 只在目标区域内解析数据
            if not in_target_area:
                continue
            
            # 跳过表头行
            if machine_name in ['机械名称', '序号']:
                continue
            
            quantity = self._get_cell_value(ws, i, 3)
            tonnage = self._get_cell_value(ws, i, 4)
            usage = self._get_cell_value(ws, i, 5)
            shift = self._get_cell_value(ws, i, 6)
            remarks = self._get_cell_value(ws, i, 7)
            
            # 只保存有机械名称的记录
            if machine_name:
                machinery.append({
                    "seqNo": seq_no,
                    "machineName": machine_name,
                    "quantity": quantity,
                    "tonnage": tonnage,
                    "usage": usage,
                    "shift": shift,
                    "remarks": remarks
                })
        return machinery
    
    def _parse_problem_feedbacks(self, ws, start_row: int, end_row: int) -> List[Dict]:
        """解析问题反馈（区域四 -> 序号1.x）"""
        problems = []
        in_target_area = False  # 是否进入目标区域（四）
        
        for i in range(start_row, end_row + 1):
            problem_no = self._get_cell_value(ws, i, 1)
            description = self._get_cell_value(ws, i, 2)
            
            # 检测区域标题
            if problem_no == '四':
                in_target_area = True
                continue  # 跳过标题行
            
            # 遇到下一个区域标题，停止解析
            if problem_no in ['五', '六'] and in_target_area:
                break
            
            # 只在目标区域内解析数据
            if not in_target_area:
                continue
            
            # 跳过表头行和子标题行
            if description in ['问题描述', '问题反馈', '序号']:
                continue
            
            # 跳过子标题行（序号为"1"且内容为"问题描述"）
            if problem_no == '1':
                continue
            
            reason = self._get_cell_value(ws, i, 4)
            impact = self._get_cell_value(ws, i, 5)
            progress = self._get_cell_value(ws, i, 6)
            
            # 保存有问题描述且序号为纯数字（2、3、4...）的记录
            if description and problem_no.isdigit():
                problems.append({
                    "problemNo": problem_no,
                    "description": description,
                    "reason": reason,
                    "impact": impact,
                    "progress": progress
                })
        return problems
    
    def _parse_requirements(self, ws, start_row: int, end_row: int) -> List[Dict]:
        """解析需求描述（区域四 -> 子区域2）"""
        requirements = []
        in_target_area = False  # 是否进入目标区域（四）
        in_requirements_section = False  # 是否进入需求描述子区域（2）
        
        for i in range(start_row, end_row + 1):
            req_no = self._get_cell_value(ws, i, 1)
            description = self._get_cell_value(ws, i, 2)
            
            # 检测区域标题
            if req_no == '四':
                in_target_area = True
                continue  # 跳过标题行
            
            # 遇到下一个区域标题，停止解析
            if req_no in ['五', '六'] and in_target_area:
                break
            
            # 只在目标区域内解析数据
            if not in_target_area:
                continue
            
            # 检测需求描述子标题（序号为"2"且内容为"需求描述"）
            if req_no == '2' and description in ['需求描述', '需求']:
                in_requirements_section = True
                continue  # 跳过子标题行
            
            # 只在需求描述子区域内解析数据
            if not in_requirements_section:
                continue
            
            # 跳过表头行
            if description in ['需求描述', '问题反馈', '序号']:
                continue
            
            urgency_level = self._get_cell_value(ws, i, 4)
            expected_time = self._get_cell_value(ws, i, 6)
            
            # 保存有需求描述的记录（任何有内容的行）
            if description:
                requirements.append({
                    "requirementNo": req_no,
                    "description": description,
                    "urgencyLevel": urgency_level,
                    "expectedTime": expected_time
                })
        return requirements
    
    def parse_all_sheets(self) -> List[Dict[str, Any]]:
        """解析所有工作表"""
        all_reports = []
        for sheet_name in self.workbook.sheetnames:
            try:
                report = self.parse_sheet(sheet_name)
                all_reports.append(report)
                print(f"✓ 成功解析工作表: {sheet_name}")
            except Exception as e:
                print(f"✗ 解析工作表 {sheet_name} 失败: {str(e)}")
        return all_reports
    
    def generate_sql_insert(self, report_data: Dict[str, Any], project_id: int, reporter_id: int) -> str:
        """
        生成SQL插入语句
        :param report_data: 报告数据
        :param project_id: 项目ID
        :param reporter_id: 填报人ID
        :return: SQL插入语句
        """
        # 转换JSON字段
        task_progress_list = json.dumps(report_data["taskProgressList"], ensure_ascii=False)
        tomorrow_plans = json.dumps(report_data["tomorrowPlans"], ensure_ascii=False)
        worker_reports = json.dumps(report_data["workerReports"], ensure_ascii=False)
        machinery_rentals = json.dumps(report_data["machineryRentals"], ensure_ascii=False)
        problem_feedbacks = json.dumps(report_data["problemFeedbacks"], ensure_ascii=False)
        requirements = json.dumps(report_data["requirements"], ensure_ascii=False)
        
        sql = f"""
INSERT INTO project_daily_reports (
    project_id, report_date, reporter_id, 
    overall_progress, progress_description,
    task_progress_list, tomorrow_plans, worker_reports,
    machinery_rentals, problem_feedbacks, requirements,
    on_site_personnel_count, status,
    created_by, updated_by, created_at, updated_at
) VALUES (
    {project_id}, 
    '{report_data['reportDate']}', 
    {reporter_id},
    '{report_data['overallProgress']}',
    '{report_data['progressDescription']}',
    '{task_progress_list}',
    '{tomorrow_plans}',
    '{worker_reports}',
    '{machinery_rentals}',
    '{problem_feedbacks}',
    '{requirements}',
    {report_data['onSitePersonnelCount']},
    'submitted',
    {reporter_id},
    {reporter_id},
    NOW(),
    NOW()
);
"""
        return sql


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("使用方法: python parse_daily_report_excel.py <excel文件路径> [输出JSON文件路径]")
        print("示例: python parse_daily_report_excel.py docs/assets/淮安日报2025.10.19.xlsx output.json")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    print(f"开始解析Excel文件: {excel_path}")
    print("=" * 80)
    
    try:
        parser = DailyReportExcelParser(excel_path)
        
        # 解析所有工作表
        all_reports = parser.parse_all_sheets()
        
        print("=" * 80)
        print(f"解析完成！共解析 {len(all_reports)} 个工作表")
        
        # 输出统计信息
        for report in all_reports:
            print(f"\n日期: {report['reportDate']}")
            print(f"  - 项目名称: {report['reporterName']}")
            print(f"  - 整体进度: {report['overallProgress']}")
            print(f"  - 进度描述: {report['progressDescription']}")
            print(f"  - 任务数量: {len(report['taskProgressList'])}")
            print(f"  - 明日计划: {len(report['tomorrowPlans'])}")
            print(f"  - 工作人员: {len(report['workerReports'])}")
            print(f"  - 现场总人数: {report['onSitePersonnelCount']}")
            print(f"  - 机械租赁: {len(report['machineryRentals'])}")
            print(f"  - 问题反馈: {len(report['problemFeedbacks'])}")
            print(f"  - 需求数量: {len(report['requirements'])}")
        
        # 保存到JSON文件
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(all_reports, f, ensure_ascii=False, indent=2)
            print(f"\n✓ 数据已保存到: {output_path}")
        else:
            # 输出第一个报告的详细JSON
            print("\n" + "=" * 80)
            print("第一个报告的详细数据（JSON格式）:")
            print("=" * 80)
            print(json.dumps(all_reports[0], ensure_ascii=False, indent=2))
        
        # 生成示例SQL
        print("\n" + "=" * 80)
        print("示例SQL插入语句（使用第一个报告）:")
        print("=" * 80)
        print(parser.generate_sql_insert(all_reports[0], project_id=1, reporter_id=1))
        
    except Exception as e:
        print(f"错误: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

